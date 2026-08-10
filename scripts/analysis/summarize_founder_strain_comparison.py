"""Reproduce the founder-strain comparison reported in Table 6."""

import csv
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
INPUTS = ROOT / "data" / "inputs"
OUTPUTS = ROOT / "data" / "outputs"


def read_rows(path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


founders = read_rows(INPUTS / "founder_strain_mpd_reed1.csv")
# Use the four values displayed in the manuscript table. Retain full precision in
# the derived output; presentation-level rounding is applied by the verifier.
founder_mean = sum(float(row["reported_weight_g"]) for row in founders) / len(founders)

itp = read_rows(INPUTS / "early_weight_change_input.csv")
itp_male_controls = [
    float(row["bw_6"])
    for row in itp
    if row["sex"] == "m" and row["is_control"].lower() == "true" and row["bw_6"]
]
itp_mean = sum(itp_male_controls) / len(itp_male_controls)

sensitivities = read_rows(OUTPUTS / "matched_quartile_standardized_survival.csv")
lowest = [
    row for row in sensitivities
    if row["quartile"] == "1"
]
if len(lowest) != 1:
    raise RuntimeError(f"Expected one lowest-quarter row; found {len(lowest)}")
lowest_mean = float(lowest[0]["mean_control_weight"])

output = [
    {"quantity": "four_founder_mean_weight_g", "value": founder_mean},
    {"quantity": "itp_male_control_6m_n", "value": len(itp_male_controls)},
    {"quantity": "itp_male_control_6m_mean_weight_g", "value": itp_mean},
    {"quantity": "itp_minus_founder_mean_g", "value": itp_mean - founder_mean},
    {"quantity": "itp_percent_above_founder_mean", "value": 100 * (itp_mean / founder_mean - 1)},
    {"quantity": "lowest_itp_quarter_mean_weight_g", "value": lowest_mean},
    {"quantity": "lowest_itp_quarter_percent_above_founder_mean",
     "value": 100 * (lowest_mean / founder_mean - 1)},
    {"quantity": "heaviest_founder_reported_mean_weight_g", "value": max(float(row["reported_weight_g"]) for row in founders)},
]

out_path = OUTPUTS / "founder_strain_weight_comparison.csv"
with out_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=["quantity", "value"])
    writer.writeheader()
    writer.writerows(output)
print(out_path)


def strain_means(path, sex):
    raw = json.loads(path.read_text(encoding="utf-8"))
    rows = raw.get("strainmeans", raw) if isinstance(raw, dict) else raw
    return {
        row["strain"]: float(row["mean"])
        for row in rows
        if str(row.get("sex", "")).lower().startswith(sex)
        and row.get("mean") is not None
    }


weights = strain_means(INPUTS / "mpd_32603_strainmeans.json", "m")
ages = strain_means(INPUTS / "mpd_32680_strainmeans.json", "m")
age_matched = {strain: value for strain, value in weights.items()
               if strain in ages and ages[strain] <= 36.0}
male_fat = strain_means(INPUTS / "mpd_10331_strainmeans.json", "m")
female_fat = strain_means(INPUTS / "mpd_10331_strainmeans.json", "f")
workbook = load_workbook(ROOT / "supplementary" / "Supplementary_Data_1-6.xlsx",
                         read_only=True, data_only=True)
sheet = workbook["SD2_body_composition"]
rows = sheet.iter_rows(values_only=True)
headers = next(rows)
body = [dict(zip(headers, row)) for row in rows]
workbook.close()
itp_male_fat = next(float(row["bodyfat_pct_control"]) for row in body
                    if row["strain"] == "UM-HET3" and row["gender"] == "Male"
                    and row["Intervention"] == "Rapamycin" and "42 ppm" in str(row["dosage"]))
itp_female_fat = next(float(row["bodyfat_pct_control"]) for row in body
                      if row["strain"] == "UM-HET3" and row["gender"] == "Female"
                      and row["Intervention"] == "Rapamycin" and "42 ppm" in str(row["dosage"]))
strain_context = [
    {"quantity": "age_matched_strains", "value": len(age_matched)},
    {"quantity": "survey_mean_weight_g",
     "value": sum(age_matched.values()) / len(age_matched)},
    {"quantity": "panel_a_strains", "value": len(age_matched)},
    {"quantity": "panel_a_at_or_above_itp",
     "value": sum(value >= lowest_mean for value in age_matched.values())},
    {"quantity": "panel_b_strains", "value": len(male_fat)},
    {"quantity": "panel_b_at_or_above_itp",
     "value": sum(value >= itp_male_fat for value in male_fat.values())},
    {"quantity": "panel_c_strains", "value": len(female_fat)},
    {"quantity": "panel_c_at_or_above_itp",
     "value": sum(value >= itp_female_fat for value in female_fat.values())},
]
strain_path = OUTPUTS / "strain_context_summary.csv"
with strain_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=["quantity", "value"])
    writer.writeheader()
    writer.writerows(strain_context)
print(strain_path)
